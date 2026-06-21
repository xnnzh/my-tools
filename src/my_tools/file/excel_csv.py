import csv
from collections.abc import Sequence
from datetime import date, datetime, time, timedelta
from io import TextIOBase

from openpyxl import load_workbook


def list_sheet_names(excel_file: str) -> list[str]:
    wb = load_workbook(excel_file, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def select_worksheet(workbook, *, sheet: str | None, sheet_index: int | None):
    if sheet is not None and sheet_index is not None:
        raise ValueError("--sheet 与 --sheet-index 不能同时指定")

    if sheet is not None:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet}' 不存在")
        return workbook[sheet]

    if sheet_index is not None:
        if sheet_index < 1:
            raise ValueError(f"sheet-index 必须大于等于 1: {sheet_index}")
        if sheet_index > len(workbook.sheetnames):
            raise ValueError(
                f"sheet-index {sheet_index} 越界，总共 {len(workbook.sheetnames)} 个工作表"
            )
        return workbook.worksheets[sheet_index - 1]

    return workbook.active


def _format_datetime(value: datetime, date_format: str) -> str:
    if date_format == "iso":
        return value.isoformat()
    return value.strftime(date_format)


def _format_date(value: date, date_format: str) -> str:
    if date_format == "iso":
        return value.isoformat()
    return value.strftime(date_format)


def _format_time(value: time, date_format: str) -> str:
    if date_format == "iso":
        return value.isoformat()
    dt = datetime.combine(date.today(), value)
    return dt.strftime(date_format)


def _format_timedelta(value: timedelta, date_format: str) -> str:
    if date_format == "iso":
        total = value.total_seconds()
        hours = int(total // 3600)
        minutes = int((total % 3600) // 60)
        seconds = total % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:06.3f}"
    return str(value)


def format_cell_value(value: object, *, empty: str = "", date_format: str = "iso") -> str:
    if value is None:
        return empty
    if isinstance(value, datetime):
        return _format_datetime(value, date_format)
    if isinstance(value, date):
        return _format_date(value, date_format)
    if isinstance(value, time):
        return _format_time(value, date_format)
    if isinstance(value, timedelta):
        return _format_timedelta(value, date_format)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value)


def trim_trailing_empty_values(row: Sequence[object], *, empty: str = "") -> list[object]:
    result = list(row)
    while result and _is_empty(result[-1], empty):
        result.pop()
    return result


def _is_empty(value: object, empty: str) -> bool:
    if empty:
        return value is None or str(value) == empty or (isinstance(value, str) and value == "")
    return value is None or (isinstance(value, str) and value == "")


def write_excel_sheet_to_csv(
    excel_file: str,
    output: TextIOBase,
    *,
    sheet: str | None = None,
    sheet_index: int | None = None,
    delimiter: str = ",",
    empty: str = "",
    date_format: str = "iso",
    data_only: bool = True,
    trim_trailing_empty_flag: bool = False,
) -> None:
    wb = load_workbook(excel_file, read_only=True, data_only=data_only)
    try:
        ws = select_worksheet(wb, sheet=sheet, sheet_index=sheet_index)
        writer = csv.writer(output, delimiter=delimiter, lineterminator="\n")
        for row in ws.iter_rows(values_only=True):
            values = [format_cell_value(v, empty=empty, date_format=date_format) for v in row]
            if trim_trailing_empty_flag:
                values = trim_trailing_empty_values(values, empty=empty)
            writer.writerow(values)
    finally:
        wb.close()
