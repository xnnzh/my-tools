from datetime import date, datetime, time
from pathlib import Path

import pytest
from click.testing import CliRunner
from openpyxl import Workbook

from my_tools.cli import cli
from my_tools.file.excel_csv import (
    format_cell_value,
    list_sheet_names,
    select_worksheet,
    trim_trailing_empty_values,
)


def _make_excel(tmp_path: Path, name: str = "test.xlsx", data: list[list] | None = None) -> Path:
    path = tmp_path / name
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if data:
        for row in data:
            ws.append(row)
    wb.save(path)
    wb.close()
    return path


def _make_multi_sheet(tmp_path: Path) -> Path:
    path = tmp_path / "multi.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "SheetA"
    ws1.append(["a", "b"])
    ws2 = wb.create_sheet("SheetB")
    ws2.append(["c", "d"])
    wb.save(path)
    wb.close()
    return path


def _read_csv_output(fp) -> str:
    fp.seek(0)
    return fp.read()


class TestListSheetNames:
    def test_basic(self):
        path = _make_multi_sheet(Path("/tmp"))
        names = list_sheet_names(str(path))
        assert names == ["SheetA", "SheetB"]


class TestSelectWorksheet:
    def test_by_name(self):
        path = _make_multi_sheet(Path("/tmp"))
        wb = Workbook(str(path) if False else None)
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        try:
            ws = select_worksheet(wb, sheet="SheetB", sheet_index=None)
            assert ws.title == "SheetB"
        finally:
            wb.close()

    def test_by_index(self):
        from openpyxl import load_workbook
        path = _make_multi_sheet(Path("/tmp"))
        wb = load_workbook(path, read_only=True)
        try:
            ws = select_worksheet(wb, sheet=None, sheet_index=2)
            assert ws.title == "SheetB"
        finally:
            wb.close()

    def test_both_raises(self):
        from openpyxl import load_workbook
        path = _make_excel(Path("/tmp"))
        wb = load_workbook(path, read_only=True)
        try:
            with pytest.raises(ValueError, match="不能同时指定"):
                select_worksheet(wb, sheet="x", sheet_index=1)
        finally:
            wb.close()

    def test_index_lt_1_raises(self):
        from openpyxl import load_workbook
        path = _make_excel(Path("/tmp"))
        wb = load_workbook(path, read_only=True)
        try:
            with pytest.raises(ValueError, match="必须大于等于 1"):
                select_worksheet(wb, sheet=None, sheet_index=0)
        finally:
            wb.close()

    def test_index_oob_raises(self):
        from openpyxl import load_workbook
        path = _make_excel(Path("/tmp"))
        wb = load_workbook(path, read_only=True)
        try:
            with pytest.raises(ValueError, match="越界"):
                select_worksheet(wb, sheet=None, sheet_index=99)
        finally:
            wb.close()

    def test_name_not_found_raises(self):
        from openpyxl import load_workbook
        path = _make_excel(Path("/tmp"))
        wb = load_workbook(path, read_only=True)
        try:
            with pytest.raises(ValueError, match="不存在"):
                select_worksheet(wb, sheet="nonexistent", sheet_index=None)
        finally:
            wb.close()

    def test_default_active(self):
        from openpyxl import load_workbook
        path = _make_excel(Path("/tmp"))
        wb = load_workbook(path, read_only=True)
        try:
            ws = select_worksheet(wb, sheet=None, sheet_index=None)
            assert ws.title == "Sheet1"
        finally:
            wb.close()


class TestFormatCellValue:
    def test_none(self):
        assert format_cell_value(None) == ""

    def test_none_with_empty(self):
        assert format_cell_value(None, empty="NULL") == "NULL"

    def test_string(self):
        assert format_cell_value("hello") == "hello"

    def test_int(self):
        assert format_cell_value(42) == "42"

    def test_float_int(self):
        assert format_cell_value(3.0) == "3"

    def test_float(self):
        assert format_cell_value(3.14) == "3.14"

    def test_bool_true(self):
        assert format_cell_value(True) == "TRUE"

    def test_bool_false(self):
        assert format_cell_value(False) == "FALSE"

    def test_date_iso(self):
        assert format_cell_value(date(2026, 6, 21)) == "2026-06-21"

    def test_datetime_iso(self):
        assert format_cell_value(datetime(2026, 6, 21, 10, 30, 0)) == "2026-06-21T10:30:00"

    def test_time_iso(self):
        assert format_cell_value(time(10, 30, 0)) == "10:30:00"

    def test_date_custom_format(self):
        assert format_cell_value(date(2026, 6, 21), date_format="%Y/%m/%d") == "2026/06/21"

    def test_datetime_custom_format(self):
        d = datetime(2026, 6, 21, 10, 30, 0)
        assert format_cell_value(d, date_format="%Y-%m-%d %H:%M") == "2026-06-21 10:30"


class TestTrimTrailingEmptyValues:
    def test_no_trailing(self):
        assert trim_trailing_empty_values(["a", "b"]) == ["a", "b"]

    def test_trailing_none(self):
        assert trim_trailing_empty_values(["a", None, None]) == ["a"]

    def test_trailing_empty_string(self):
        assert trim_trailing_empty_values(["a", "", ""]) == ["a"]

    def test_mixed(self):
        assert trim_trailing_empty_values(["a", "", "b"]) == ["a", "", "b"]


class TestConvertBasic:
    def test_stdout(self):
        path = _make_excel(Path("/tmp"))
        result = CliRunner().invoke(cli, ["file", "excel-to-csv", str(path)])
        assert result.exit_code == 0
        assert result.output == ""

    def test_simple_data(self):
        path = _make_excel(Path("/tmp"), data=[["a", "b"], ["1", "2"]])
        result = CliRunner().invoke(cli, ["file", "excel-to-csv", str(path)])
        assert result.exit_code == 0
        assert result.output == "a,b\n1,2\n"

    def test_to_file(self, tmp_path):
        path = _make_excel(tmp_path, data=[["a", "b"], ["1", "2"]])
        out = tmp_path / "out.csv"
        result = CliRunner().invoke(cli, ["file", "excel-to-csv", str(path), "-o", str(out)])
        assert result.exit_code == 0
        assert out.read_text(encoding="utf-8") == "a,b\n1,2\n"

    def test_list_sheets(self):
        path = _make_multi_sheet(Path("/tmp"))
        result = CliRunner().invoke(cli, ["file", "excel-to-csv", str(path), "--list-sheets"])
        assert result.exit_code == 0
        assert result.output == "SheetA\nSheetB\n"

    def test_sheet_by_name(self):
        path = _make_multi_sheet(Path("/tmp"))
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--sheet", "SheetB"]
        )
        assert result.exit_code == 0
        assert result.output == "c,d\n"

    def test_sheet_by_index(self):
        path = _make_multi_sheet(Path("/tmp"))
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--sheet-index", "2"]
        )
        assert result.exit_code == 0
        assert result.output == "c,d\n"

    def test_delimiter_semicolon(self):
        path = _make_excel(Path("/tmp"), data=[["a", "b"], ["1", "2"]])
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--delimiter", ";"]
        )
        assert result.exit_code == 0
        assert result.output == "a;b\n1;2\n"

    def test_empty_cells_default(self):
        path = _make_excel(Path("/tmp"), data=[["a", None], [None, "b"]])
        result = CliRunner().invoke(cli, ["file", "excel-to-csv", str(path)])
        assert result.exit_code == 0
        assert result.output == "a,\n,b\n"

    def test_empty_cells_with_flag(self):
        path = _make_excel(Path("/tmp"), data=[["a", None], [None, "b"]])
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--empty", "NULL"]
        )
        assert result.exit_code == 0
        assert result.output == "a,NULL\nNULL,b\n"

    def test_trim_trailing_empty(self):
        path = _make_excel(Path("/tmp"), data=[["a", "b", None], ["c", None, None]])
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--trim-trailing-empty"]
        )
        assert result.exit_code == 0
        assert result.output == "a,b\nc\n"

    def test_sheet_not_found(self):
        path = _make_excel(Path("/tmp"))
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--sheet", "nonexistent"]
        )
        assert result.exit_code != 0
        assert "不存在" in result.output

    def test_sheet_index_oob(self):
        path = _make_excel(Path("/tmp"))
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--sheet-index", "99"]
        )
        assert result.exit_code != 0
        assert "越界" in result.output

    def test_sheet_and_index_both(self):
        path = _make_excel(Path("/tmp"))
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--sheet", "Sheet1", "--sheet-index", "1"]
        )
        assert result.exit_code != 0
        assert "不能同时指定" in result.output

    def test_help(self):
        result = CliRunner().invoke(cli, ["file", "excel-to-csv", "--help"])
        assert result.exit_code == 0
        assert "--sheet" in result.output
        assert "--sheet-index" in result.output
        assert "--list-sheets" in result.output
        assert "--delimiter" in result.output
        assert "--encoding" in result.output
        assert "--empty" in result.output
        assert "--date-format" in result.output
        assert "--formula" in result.output
        assert "--trim-trailing-empty" in result.output

    def test_encoding_gbk_output(self, tmp_path):
        path = _make_excel(tmp_path, data=[["a"], ["你好"]])
        out = tmp_path / "out.csv"
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "-o", str(out), "--encoding", "gbk"]
        )
        assert result.exit_code == 0
        content = out.read_bytes()
        assert content == "a\n你好\n".encode("gbk")


class TestConvertDates:
    def test_date_iso_default(self, tmp_path):
        path = _make_excel(tmp_path, data=[[date(2026, 6, 21)]])
        result = CliRunner().invoke(cli, ["file", "excel-to-csv", str(path)])
        assert result.exit_code == 0
        assert result.output == "2026-06-21T00:00:00\n"

    def test_date_custom_format(self, tmp_path):
        path = _make_excel(tmp_path, data=[[date(2026, 6, 21)]])
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--date-format", "%Y/%m/%d"]
        )
        assert result.exit_code == 0
        assert result.output == "2026/06/21\n"

    def test_datetime_iso(self, tmp_path):
        path = _make_excel(tmp_path, data=[[datetime(2026, 6, 21, 10, 30, 0)]])
        result = CliRunner().invoke(cli, ["file", "excel-to-csv", str(path)])
        assert result.exit_code == 0
        assert result.output == "2026-06-21T10:30:00\n"

    def test_formula_cache(self, tmp_path):
        path = tmp_path / "formula.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        ws["A2"] = 2
        ws["A3"] = "=SUM(A1:A2)"
        wb.save(path)
        wb.close()
        result = CliRunner().invoke(cli, ["file", "excel-to-csv", str(path)])
        assert result.exit_code == 0
        assert "=SUM" not in result.output

    def test_formula_text(self, tmp_path):
        path = tmp_path / "formula.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        ws["A2"] = 2
        ws["A3"] = "=SUM(A1:A2)"
        wb.save(path)
        wb.close()
        result = CliRunner().invoke(
            cli, ["file", "excel-to-csv", str(path), "--formula"]
        )
        assert result.exit_code == 0
        assert "=SUM(A1:A2)" in result.output
